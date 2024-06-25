from flask import Flask, render_template, request, jsonify
import openpyxl
import datetime
import os

app = Flask(__name__)

# Set paths
wb_path = os.path.join(os.path.dirname(__file__), 'Attendance.xlsx')
source_path = os.path.join(os.path.dirname(__file__), 'Datasheet.xlsx')

# Create and save the workbook if it doesn't exist
try:
    wb_attendance = openpyxl.load_workbook(wb_path)
except FileNotFoundError:
    wb_attendance = openpyxl.Workbook()
    wb_attendance.save(wb_path)
    wb_attendance.close()

# Load source workbook
wb_source = openpyxl.load_workbook(source_path)
source_sheet = wb_source['Data']

@app.route('/')
def index():
    return render_template('index.html', people=get_names())

@app.route('/submit', methods=['POST'])
def submit():
    data = request.json
    attendance_array = data.get('attendance', [])

    # Set up paths again (in case new file created)
    wb_path = os.path.join(os.path.dirname(__file__), 'Attendance.xlsx')

    # Load attendance workbook
    wb_attendance = openpyxl.load_workbook(wb_path)
    today = datetime.datetime.now().strftime("%d-%m-%y")
    new_sheet = get_or_create_sheet(wb_attendance, today)

    # Update attendance data
    for i, attendance in enumerate(attendance_array, start=2):
        new_sheet.cell(row=i, column=2, value=attendance['value'])

    # Save the attendance workbook
    wb_attendance.save(wb_path)
    return jsonify({"status": "success", "attendance_array": attendance_array})

def get_names():
    # Update total names and details
    total_details = source_sheet.max_column
    total_names = source_sheet.max_row

    # Copy names
    name_array = []
    for row in range(2, total_names + 1):
        name_array.append(source_sheet.cell(row=row, column=1).value)

    return name_array

def get_or_create_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        new_sheet = workbook.create_sheet(sheet_name)
        total_names = len(get_names())

        # Copy names to new sheet
        for row in range(2, total_names + 2):
            source_cell = source_sheet.cell(row=row, column=1)
            dest_cell = new_sheet.cell(row=row, column=1)
            dest_cell.value = source_cell.value
    else:
        new_sheet = workbook[sheet_name]

    return new_sheet

if __name__ == '__main__':
    app.run(debug=True)
